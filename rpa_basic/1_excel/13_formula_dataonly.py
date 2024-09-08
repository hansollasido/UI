from openpyxl import load_workbook
# wb = load_workbook("sample_formula.xlsx")
# ws = wb.active

wb = load_workbook("sample_formula.xlsx", data_only=True)
ws = wb.active

# 수식이 아닌 실제 데이터만 가져옴
# evaluate 되지 않은 상태의 데이터 none 이라고 표시

### => None을 데이터로 저장하고 싶으면, 한번 엑셀 파일을 열었다가 저장하면 됨
for row in ws.values:
    for cell in row:
        print(cell)